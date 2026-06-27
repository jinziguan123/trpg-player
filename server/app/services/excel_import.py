"""从 COC 七版半自动人物卡 Excel 导入角色数据

支持 v1.8.6 模板及常见变体布局。
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl


ATTR_CELLS = {
    "STR": "K3",
    "DEX": "N3",
    "POW": "Q3",
    "CON": "K5",
    "APP": "N5",
    "EDU": "Q5",
    "SIZ": "K7",
    "INT": "N7",
}

SKILL_ROWS_LEFT = list(range(16, 49))
SKILL_ROWS_RIGHT = list(range(16, 45))

BACKSTORY_CELLS_V1 = {
    "personalDescription": "N60",
    "ideologyBeliefs": "N62",
    "significantPeople": "N64",
    "meaningfulLocations": "N66",
    "treasuredPossessions": "N68",
    "traits": "N70",
}

BACKSTORY_CELLS_V2 = {
    "personalDescription": "O61",
    "ideologyBeliefs": "O63",
    "significantPeople": "O65",
    "meaningfulLocations": "O67",
    "treasuredPossessions": "O69",
    "traits": "O71",
}

EXTRA_BACKSTORY_V1 = {
    "scarsAndWounds": "N72",
    "phobiasAndManias": "N74",
}

EXTRA_BACKSTORY_V2 = {
    "scarsAndWounds": "O75",
    "phobiasAndManias": "O77",
}

WEAPON_ROWS = list(range(53, 58))


def _cell_val(ws, coord, default=None):
    v = ws[coord].value
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    return v


def _cell_int(ws, coord, default: int = 0) -> int:
    v = ws[coord].value
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _read_text(ws, coord: str) -> str:
    v = ws[coord].value
    if v is not None and isinstance(v, str) and v.strip():
        return v.strip()
    return ""


def _calc_damage_bonus(combined: int) -> str:
    if combined <= 64:
        return "-2"
    if combined <= 84:
        return "-1"
    if combined <= 124:
        return "0"
    if combined <= 164:
        return "1D4"
    return "1D6"


def _calc_build(combined: int) -> int:
    if combined <= 64:
        return -2
    if combined <= 84:
        return -1
    if combined <= 124:
        return 0
    if combined <= 164:
        return 1
    return 2


def parse_coc_character_sheet(file_bytes: bytes) -> dict[str, Any]:
    """解析 COC 七版角色卡 Excel，返回可直接用于创建角色的数据。"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    if "人物卡" not in wb.sheetnames:
        raise ValueError("Excel 中未找到「人物卡」工作表")

    ws = wb["人物卡"]

    name = _cell_val(ws, "C3", "") or _cell_val(ws, "D3", "")
    if not name:
        raise ValueError("角色姓名为空（C3 单元格）")

    age = _cell_int(ws, "C6") or _cell_int(ws, "D6", 25)
    gender = _cell_val(ws, "F6", "") or _cell_val(ws, "G6", "")
    residence = _cell_val(ws, "C7", "") or _cell_val(ws, "D7", "")
    birthplace = _cell_val(ws, "C8", "") or _cell_val(ws, "D8", "")
    occupation = _cell_val(ws, "C5", "") or _cell_val(ws, "D5", "")

    base_attributes: dict[str, int] = {}
    for attr_key, cell in ATTR_CELLS.items():
        val = _cell_int(ws, cell)
        if val > 0:
            base_attributes[attr_key] = val

    if not base_attributes:
        raise ValueError("属性值全部为空，请确认 Excel 文件已填写")

    # ---- 派生数值 ----
    hp_current = _cell_int(ws, "D10")
    hp_max = _cell_int(ws, "E10")
    san_current = _cell_int(ws, "H10")
    san_max = _cell_int(ws, "I10", 99)
    if not san_current:
        san_current = san_max or base_attributes.get("POW", 0)
    mp_current = _cell_int(ws, "Q10")
    mp_max = _cell_int(ws, "R10")
    luck = _cell_int(ws, "M10") or _cell_int(ws, "N10")
    mov = _cell_int(ws, "Q7") or _cell_int(ws, "P7", 8)

    combined = base_attributes.get("STR", 50) + base_attributes.get("SIZ", 50)
    damage_bonus = _cell_val(ws, "V51")
    if damage_bonus is None:
        damage_bonus = _calc_damage_bonus(combined)
    build = _cell_int(ws, "V53")
    if build == 0 and combined != 0:
        build = _calc_build(combined)
    dodge = _cell_int(ws, "V55")
    if dodge == 0:
        dodge = base_attributes.get("DEX", 50) // 2

    # ---- 技能（总值在 K 列和 V 列）----
    skills: dict[str, int] = {}

    for row in SKILL_ROWS_LEFT:
        skill_name = _cell_val(ws, f"D{row}")
        if not skill_name or not isinstance(skill_name, str):
            continue
        raw = skill_name.strip()
        if raw in ("技能名称",):
            continue
        skill_name = raw.rstrip(" Ω：:")
        sub = _cell_val(ws, f"E{row}")
        if sub and isinstance(sub, str) and sub.strip():
            s = sub.strip()
            if not s.startswith("←"):
                combined_name = skill_name.rstrip("：:①②③")
                combined_name += f"({s})"
                skill_name = combined_name
        val = _cell_int(ws, f"K{row}")
        if val > 0 and skill_name != "克苏鲁神话":
            skills[skill_name] = val

    for row in SKILL_ROWS_RIGHT:
        skill_name = _cell_val(ws, f"P{row}")
        if not skill_name or not isinstance(skill_name, str):
            continue
        raw = skill_name.strip()
        if raw in ("技能名称",):
            continue
        skill_name = raw.rstrip(" Ω：:")
        sub = _cell_val(ws, f"Q{row}")
        if sub and isinstance(sub, str) and sub.strip() and sub.strip() != "可自设":
            s = sub.strip()
            if not s.startswith("←"):
                combined_name = skill_name.rstrip("：:①②③")
                combined_name += f"({s})"
                skill_name = combined_name
        val = _cell_int(ws, f"V{row}")
        if val > 0:
            skills[skill_name] = val

    # 信用评级：在左侧技能列中查找
    credit_rating = 0
    for row in SKILL_ROWS_LEFT:
        if _cell_val(ws, f"D{row}") == "信用评级":
            credit_rating = _cell_int(ws, f"K{row}")
            break
    if credit_rating > 0:
        skills["信用评级"] = credit_rating
    if dodge > 0:
        skills["闪避"] = dodge

    # ---- system_data ----
    system_data: dict[str, Any] = {
        "age": age,
        "move": mov,
        "damageBonus": str(damage_bonus) if damage_bonus else "0",
        "build": build,
    }
    if hp_max > 0:
        system_data["hitPoints"] = {"current": hp_current or hp_max, "max": hp_max}
    if san_current > 0 or san_max > 0:
        system_data["sanity"] = {"current": san_current, "max": san_max}
    if mp_max > 0:
        system_data["magicPoints"] = {"current": mp_current or mp_max, "max": mp_max}
    if luck > 0:
        system_data["luck"] = luck
    if occupation:
        system_data["occupation"] = str(occupation)
    if gender:
        system_data["gender"] = str(gender)
    if residence:
        system_data["residence"] = str(residence)
    if birthplace:
        system_data["birthplace"] = str(birthplace)

    # ---- 背景故事（支持两种布局）----
    backstory_parts: list[str] = []
    label_map = {
        "personalDescription": "个人描述",
        "ideologyBeliefs": "思想/信念",
        "significantPeople": "重要之人",
        "meaningfulLocations": "意义非凡之地",
        "treasuredPossessions": "宝贵之物",
        "traits": "特点",
    }

    bs_cells = BACKSTORY_CELLS_V1
    extra_cells = EXTRA_BACKSTORY_V1
    has_v1 = any(_read_text(ws, c) for c in BACKSTORY_CELLS_V1.values())
    if not has_v1:
        has_v2 = any(_read_text(ws, c) for c in BACKSTORY_CELLS_V2.values())
        if has_v2:
            bs_cells = BACKSTORY_CELLS_V2
            extra_cells = EXTRA_BACKSTORY_V2

    for key, cell in bs_cells.items():
        val = _read_text(ws, cell)
        if val:
            system_data[key] = val
            backstory_parts.append(f"【{label_map.get(key, key)}】{val}")

    extra_label = {"scarsAndWounds": "伤口/疤痕", "phobiasAndManias": "恐惧症/狂躁症"}
    for key, cell in extra_cells.items():
        val = _read_text(ws, cell)
        if val and val != "暂无":
            system_data[key] = val
            backstory_parts.append(f"【{extra_label.get(key, key)}】{val}")

    for free_cell in ("L76", "M79"):
        free_backstory = _read_text(ws, free_cell)
        if free_backstory and free_backstory != "请务必在此填写背景故事！":
            backstory_parts.append(free_backstory)
            break

    backstory = "\n".join(backstory_parts)

    # ---- 武器 ----
    weapons: list[dict[str, Any]] = []
    for row in WEAPON_ROWS:
        wname = _cell_val(ws, f"B{row}")
        if not wname or str(wname).strip() in ("", "无"):
            continue
        weapons.append({
            "name": str(wname).strip(),
            "skill": str(_cell_val(ws, f"E{row}", "")).strip(),
            "damage": str(_cell_val(ws, f"J{row}", "")).strip(),
            "range": str(_cell_val(ws, f"L{row}", "")).strip(),
            "attacks": _cell_int(ws, f"O{row}", 1),
            "ammo": str(_cell_val(ws, f"Q{row}", "")).strip(),
        })
    if weapons:
        system_data["weapons"] = weapons

    # ---- 调查员经历 ----
    history_parts: list[str] = []
    for row in range(79, 86):
        val = _read_text(ws, f"B{row}")
        if val and not val.startswith("示例") and val != "状态":
            history_parts.append(val)
    if history_parts:
        system_data["investigatorHistory"] = "\n".join(history_parts)

    wb.close()

    return {
        "name": str(name).strip(),
        "age": age,
        "base_attributes": base_attributes,
        "skills": skills,
        "system_data": system_data,
        "backstory": backstory,
        "weapons": weapons,
    }
